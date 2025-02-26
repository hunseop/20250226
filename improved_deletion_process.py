#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
방화벽 정책 관리 프로세스 개선 스크립트
기존 deletion_process.py의 로직을 유지하면서 코드 구조만 개선
"""

import pandas as pd
pd.options.mode.chained_assignment = None
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
import logging
import re
from datetime import datetime, timedelta
import os
import json
import sys

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ConfigManager:
    """설정 파일을 관리하는 클래스"""
    
    def __init__(self, config_file='config.json'):
        """
        설정 파일을 로드합니다.
        
        Args:
            config_file (str): 설정 파일 경로
        """
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            logger.info(f"설정 파일 '{config_file}'을 성공적으로 로드했습니다.")
        except FileNotFoundError:
            logger.error(f"설정 파일 '{config_file}'을 찾을 수 없습니다.")
            sys.exit(1)
        except json.JSONDecodeError:
            logger.error(f"설정 파일 '{config_file}'의 형식이 올바르지 않습니다.")
            sys.exit(1)
    
    def get(self, key, default=None):
        """
        설정값을 가져옵니다.
        
        Args:
            key (str): 설정 키
            default: 기본값
            
        Returns:
            설정값 또는 기본값
        """
        keys = key.split('.')
        value = self.config
        
        try:
            for k in keys:
                value = value[k]
            return value
        except (KeyError, TypeError):
            logger.warning(f"설정 키 '{key}'를 찾을 수 없습니다. 기본값 '{default}'를 사용합니다.")
            return default

class FileManager:
    """파일 관리 기능을 제공하는 클래스"""
    
    def __init__(self, config_manager):
        """
        파일 관리자를 초기화합니다.
        
        Args:
            config_manager (ConfigManager): 설정 관리자
        """
        self.config = config_manager
    
    def update_version(self, filename, final_version=False):
        """
        파일 이름의 버전을 업데이트합니다.
        
        Args:
            filename (str): 파일 이름
            final_version (bool): 최종 버전 여부
            
        Returns:
            str: 업데이트된 파일 이름
        """
        base_name, ext = filename.rsplit('.', 1)
        
        version_format = self.config.get('file_naming.policy_version_format', '_v{version}')
        final_suffix = self.config.get('file_naming.final_version_suffix', '_vf')
        
        match = re.search(r'_v(\d+)$', base_name)
        final_match = re.search(r'_vf$', base_name)
        
        if final_match:
            return filename
        
        if final_version:
            if match:
                new_base_name = re.sub(r'_v\d+$', final_suffix, base_name)
            else:
                new_base_name = f"{base_name}{final_suffix}"
        else:
            if match:
                version = int(match.group(1))
                new_version = version + 1
                new_base_name = re.sub(r'_v\d+$', version_format.format(version=new_version), base_name)
            else:
                new_base_name = f"{base_name}{version_format.format(version=1)}"
        
        new_filename = f"{new_base_name}.{ext}"
        return new_filename
    
    def select_files(self, extension=None):
        """
        지정된 확장자의 파일 목록에서 파일을 선택합니다.
        
        Args:
            extension (str): 파일 확장자
            
        Returns:
            str: 선택된 파일 이름 또는 None
        """
        if extension is None:
            extension = self.config.get('file_extensions.excel', '.xlsx')
            
        file_list = [file for file in os.listdir() if file.endswith(extension)]
        if not file_list:
            print(f"{extension} 확장자를 가진 파일이 없습니다.")
            return None
        
        for i, file in enumerate(file_list, start=1):
            print(f"{i}. {file}")
        
        while True:
            choice = input("파일 번호를 입력하세요 (종료: 0): ")
            if choice.isdigit():
                choice = int(choice)
                if choice == 0:
                    print('프로그램을 종료합니다.')
                    return None
                elif 1 <= choice <= len(file_list):
                    return file_list[choice - 1]
            print('유효하지 않은 번호입니다. 다시 시도하세요.')
    
    def remove_extension(self, filename):
        """
        파일 이름에서 확장자를 제거합니다.
        
        Args:
            filename (str): 파일 이름
            
        Returns:
            str: 확장자가 제거된 파일 이름
        """
        return os.path.splitext(filename)[0]

class ExcelManager:
    """Excel 파일 관리 기능을 제공하는 클래스"""
    
    def __init__(self, config_manager):
        """
        Excel 관리자를 초기화합니다.
        
        Args:
            config_manager (ConfigManager): 설정 관리자
        """
        self.config = config_manager
    
    def save_to_excel(self, df, sheet_type, file_name):
        """
        DataFrame을 Excel 파일에 저장합니다.
        
        Args:
            df (DataFrame): 저장할 DataFrame
            sheet_type (str): 시트 유형
            file_name (str): 파일 이름
        """
        wb = load_workbook(file_name)
        sheet = wb[sheet_type]
        
        # 첫 번째 행 삽입
        sheet.insert_rows(1)
        sheet['A1'] = '="대상 정책 수: "&COUNTA(B:B)-1'
        sheet['A1'].font = Font(bold=True)
        
        # 헤더 스타일 설정
        header_color = self.config.get('excel_styles.header_fill_color', 'E0E0E0')
        history_color = self.config.get('excel_styles.history_fill_color', 'ccffff')
        
        for col in range(1, 8):
            cell = sheet.cell(row=2, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        
        if sheet_type != '이력없음_미사용정책':
            for col in range(8, 24):
                cell = sheet.cell(row=2, column=col)
                cell.fill = PatternFill(start_color=history_color, end_color=history_color, fill_type='solid')
        
        wb.save(file_name)
        logger.info(f"Excel 파일 '{file_name}'에 데이터를 저장했습니다.")

class RequestParser:
    """신청 정보 파싱 기능을 제공하는 클래스"""
    
    def __init__(self, config_manager):
        """
        신청 정보 파서를 초기화합니다.
        
        Args:
            config_manager (ConfigManager): 설정 관리자
        """
        self.config = config_manager
    
    def convert_to_date(self, date_str):
        """
        날짜 문자열을 날짜 형식으로 변환합니다.
        
        Args:
            date_str (str): 날짜 문자열
            
        Returns:
            str: 변환된 날짜 문자열
        """
        try:
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            return date_obj.strftime('%Y-%m-%d')
        except ValueError:
            return date_str
    
    def parse_request_info(self, rulename, description):
        """
        규칙 이름과 설명에서 신청 정보를 파싱합니다.
        
        Args:
            rulename (str): 규칙 이름
            description (str): 설명
            
        Returns:
            dict: 파싱된 신청 정보
        """
        data_dict = {
            "Request Type": "Unknown",
            "Request ID": None,
            "Ruleset ID": None,
            "MIS ID": None,
            "Request User": None,
            "Start Date": self.convert_to_date('19000101'),
            "End Date": self.convert_to_date('19000101'),
        }
        
        if pd.isnull(description):
            return data_dict
        
        # 패턴 정의
        pattern_3 = re.compile(self.config.get('patterns.pattern_3', 'MASKED'))
        pattern_1_rulename = re.compile(self.config.get('patterns.pattern_1_rulename', 'MASKED'))
        pattern_1_user = self.config.get('patterns.pattern_1_user', 'MASKED')
        rulename_1_rulename = self.config.get('patterns.rulename_1_rulename', 'MASKED')
        rulename_1_date = self.config.get('patterns.rulename_1_date', 'MASKED')
        
        # 매칭
        match_3 = pattern_3.match(description)
        name_match = pattern_1_rulename.match(str(rulename))
        user_match = re.search(pattern_1_user, description)
        desc_match = re.search(rulename_1_rulename, description)
        date_match = re.search(rulename_1_date, description)
        
        if match_3:
            data_dict = {
                "Request Type": None,
                "Request ID": match_3.group(5),
                "Ruleset ID": match_3.group(1),
                "MIS ID": match_3.group(6) if match_3.group(6) else None,
                "Request User": match_3.group(4),
                "Start Date": self.convert_to_date(match_3.group(2)),
                "End Date": self.convert_to_date(match_3.group(3)),
            }
            
            type_code = data_dict["Request ID"][:1]
            if type_code == "P":
                data_dict["Request Type"] = "GROUP"
            elif type_code == "F":
                data_dict["Request Type"] = "NORMAL"
            elif type_code == "S":
                data_dict["Request Type"] = "SERVER"
            elif type_code == "M":
                data_dict["Request Type"] = "PAM"
            else:
                data_dict["Request Type"] = "Unknown"
        
        if name_match:
            data_dict['Request Type'] = "OLD"
            data_dict['Request ID'] = name_match.group(1)
            if user_match:
                data_dict['Request User'] = user_match.group(1).replace("*ACL*", "")
            if date_match:
                data_dict['Start Date'] = self.convert_to_date(date_match.group().split("~")[0])
                data_dict['End Date'] = self.convert_to_date(date_match.group().split("~")[1])
        
        if desc_match:
            date = description.split(';')[0]
            start_date = date.split('~')[0].replace('[', '').replace('-', '')
            end_date = date.split('~')[1].replace(']', '').replace('-', '')
            
            data_dict = {
                "Request Type": "OLD",
                "Request ID": desc_match.group(1).split('-')[1],
                "Ruleset ID": None,
                "MIS ID": None,
                "Request User": user_match.group(1).replace("*ACL*", "") if user_match else None,
                "Start Date": self.convert_to_date(start_date),
                "End Date": self.convert_to_date(end_date),
            }
        
        return data_dict
    
    def parse_request_type(self, file_manager):
        """
        파일에서 신청 유형을 파싱합니다.
        
        Args:
            file_manager (FileManager): 파일 관리자
            
        Returns:
            bool: 성공 여부
        """
        try:
            file_name = file_manager.select_files()
            if not file_name:
                return False
            
            df = pd.read_excel(file_name)
            
            for index, row in df.iterrows():
                result = self.parse_request_info(row['Rule Name'], row['Description'])
                for key, value in result.items():
                    df.at[index, key] = value
            
            new_file_name = file_manager.update_version(file_name)
            df.to_excel(new_file_name, index=False)
            logger.info(f"신청 유형 파싱 결과를 '{new_file_name}'에 저장했습니다.")
            return True
        except Exception as e:
            logger.exception(f"신청 유형 파싱 중 오류 발생: {e}")
            return False

class RequestExtractor:
    """신청 ID 추출 기능을 제공하는 클래스"""
    
    def __init__(self, config_manager):
        """
        신청 ID 추출기를 초기화합니다.
        
        Args:
            config_manager (ConfigManager): 설정 관리자
        """
        self.config = config_manager
    
    def extract_request_id(self, file_manager):
        """
        파일에서 신청 ID를 추출합니다.
        
        Args:
            file_manager (FileManager): 파일 관리자
            
        Returns:
            bool: 성공 여부
        """
        try:
            file_name = file_manager.select_files()
            if not file_name:
                return False
            
            df = pd.read_excel(file_name)
            
            # 'Unknown' 값을 제외하고 고유한 Request Type 값을 추출
            unique_types = df[df['Request Type'] != 'Unknown']['Request Type'].unique()
            
            # 고유한 Request Type 값을 최대 5개 선택
            selected_types = unique_types[:5]
            
            # 선택된 Request Type에 해당하는 데이터 추출
            selected_data = df[df['Request Type'].isin(selected_types)]
            
            # 각 Request Type별로 Request ID 값만 추출하여 중복 제거 후 Excel의 각 시트로 저장
            request_id_prefix = self.config.get('file_naming.request_id_prefix', 'request_id_')
            output_file = f"{request_id_prefix}{file_name}"
            
            with pd.ExcelWriter(output_file) as writer:
                for request_type, group in selected_data.groupby('Request Type'):
                    group[['Request ID']].drop_duplicates().to_excel(writer, sheet_name=request_type, index=False)
            
            logger.info(f"신청 ID 추출 결과를 '{output_file}'에 저장했습니다.")
            return True
        except Exception as e:
            logger.exception(f"신청 ID 추출 중 오류 발생: {e}")
            return False

class RequestInfoAdder:
    """신청 정보 추가 기능을 제공하는 클래스"""
    
    def __init__(self, config_manager):
        """
        신청 정보 추가기를 초기화합니다.
        
        Args:
            config_manager (ConfigManager): 설정 관리자
        """
        self.config = config_manager
    
    def read_and_process_excel(self, file):
        """
        Excel 파일을 읽고 초기 처리합니다.
        
        Args:
            file (str): 파일 경로
            
        Returns:
            DataFrame: 처리된 DataFrame
        """
        df = pd.read_excel(file)
        df.replace({'nan': None}, inplace=True)
        return df.astype(str)
    
    def match_and_update_df(self, rule_df, info_df):
        """
        조건에 따라 DataFrame의 값을 매칭 및 업데이트합니다.
        
        Args:
            rule_df (DataFrame): 규칙 DataFrame
            info_df (DataFrame): 정보 DataFrame
        """
        total = len(rule_df)
        for idx, row in rule_df.iterrows():
            print(f"\r진행 상황: {idx + 1}/{total}", end='', flush=True)
            if row['Request Type'] == 'GROUP':
                matched_row = info_df[
                    ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['MIS_ID'] == row['MIS ID'])) |
                    ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['REQUEST_END_DATE'] == row['End Date']) & (info_df['WRITE_PERSON_ID'] == row['Request User'])) |
                    ((info_df['REQUEST_ID'] == row['Request ID']) & (info_df['REQUEST_END_DATE'] == row['End Date']) & (info_df['REQUESTER_ID'] == row['Request User']))
                ]
            else:
                matched_row = info_df[info_df['REQUEST_ID'] == row['Request ID']]
            
            if not matched_row.empty:
                for col in matched_row.columns:
                    if col in ['REQUEST_START_DATE', 'REQUEST_END_DATE', 'Start Date', 'End Date']:
                        rule_df.at[idx, col] = pd.to_datetime(matched_row[col].values[0], errors='coerce')
                    else:
                        rule_df.at[idx, col] = matched_row[col].values[0]
            elif row['Request Type'] != 'nan' and row['Request Type'] != 'Unknown':
                rule_df.at[idx, 'REQUEST_ID'] = row['Request ID']
                rule_df.at[idx, 'REQUEST_START_DATE'] = row['Start Date']
                rule_df.at[idx, 'REQUEST_END_DATE'] = row['End Date']
                rule_df.at[idx, 'REQUESTER_ID'] = row['Request User']
                rule_df.at[idx, 'REQUESTER_EMAIL'] = row['Request User'] + '@gmail.com'
    
    def find_auto_extension_id(self, file_manager):
        """
        자동 연장 ID를 찾습니다.
        
        Args:
            file_manager (FileManager): 파일 관리자
            
        Returns:
            Series: 자동 연장 ID 시리즈
        """
        print('가공된 신청정보 파일을 선택하세요.')
        selected_file = file_manager.select_files()
        if not selected_file:
            return pd.Series()
        
        df = pd.read_excel(selected_file)
        filtered_df = df[df['REQEUST_STATUS'].isin([98, 99])]['REQUEST_ID'].drop_duplicates()
        
        return filtered_df
    
    def add_request_info(self, file_manager):
        """
        파일에 신청 정보를 추가합니다.
        
        Args:
            file_manager (FileManager): 파일 관리자
            
        Returns:
            bool: 성공 여부
        """
        try:
            print('정책 파일을 선택하세요:')
            rule_file = file_manager.select_files()
            if not rule_file:
                return False
            
            print("정보 파일을 선택하세요:")
            info_file = file_manager.select_files()
            if not info_file:
                return False
            
            rule_df = self.read_and_process_excel(rule_file)
            info_df = self.read_and_process_excel(info_file)
            info_df = info_df.sort_values(by='REQUEST_END_DATE', ascending=False)
            
            auto_extension_id = self.find_auto_extension_id(file_manager)
            
            self.match_and_update_df(rule_df, info_df)
            rule_df.replace({'nan': None}, inplace=True)
            
            rule_df.loc[rule_df['REQUEST_ID'].isin(auto_extension_id), 'REQUEST_STATUS'] = '99'
            
            new_file_name = file_manager.update_version(rule_file)
            rule_df.to_excel(new_file_name, index=False)
            logger.info(f"신청 정보 추가 결과를 '{new_file_name}'에 저장했습니다.")
            return True
        except Exception as e:
            logger.exception(f"신청 정보 추가 중 오류 발생: {e}")
            return False

# 메인 함수와 기타 함수들은 다음 단계에서 추가할 예정입니다.

if __name__ == "__main__":
    print("방화벽 정책 관리 프로세스 개선 스크립트")
    print("개발 중입니다. 아직 실행할 수 없습니다.")
